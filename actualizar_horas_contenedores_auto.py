import pandas as pd
import os
from datetime import datetime, timedelta, time
import re
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def find_header_row(file_path, sheet_name, essential_headers, max_rows_to_check=20):
    """
    Finds the 0-indexed row number containing all essential_headers.
    Searches within the first `max_rows_to_check` of the specified sheet.
    Returns the 0-indexed row number if found, otherwise raises ValueError.
    """
    try:
        df_peek = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=max_rows_to_check)
    except Exception as e:
        try:
            df_peek = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=max_rows_to_check, engine='openpyxl')
        except Exception as e_openpyxl:
            raise ValueError(f"Could not read sheet '{sheet_name}' from '{file_path}' to find headers. Default engine error: {e}. Openpyxl engine error: {e_openpyxl}")

    normalized_essential_headers = {str(h).strip().lower() for h in essential_headers}
    if not normalized_essential_headers:
        raise ValueError("Essential headers list cannot be empty.")

    for i, row in df_peek.iterrows():
        row_values = {str(cell).strip().lower() for cell in row.tolist() if pd.notna(cell) and str(cell).strip()}
        if normalized_essential_headers.issubset(row_values):
            print(f"Found all essential headers ({', '.join(essential_headers)}) in row {i+1} (0-indexed: {i}) of sheet '{sheet_name}'.")
            return i 

    raise ValueError(
        f"Could not find all essential headers ({', '.join(essential_headers)}) "
        f"in the first {max_rows_to_check} rows of sheet '{sheet_name}' in file '{file_path}'. "
        f"Please ensure these columns exist and are within the first {max_rows_to_check} rows."
    )

def update_export_times(main_file_path, export_log_path, sheet_date=None):
    """
    Updates the 'Hr Salida QP' column in the main file with times from the export container log
    by matching container numbers and license plates.
    """
    updates_count = 0
    
    if sheet_date is None:
        yesterday = datetime.now() - timedelta(days=1)
        sheet_date = yesterday.strftime('%d')
    
    print(f"Loading export log data from sheet: {sheet_date}")
    
    try:
        dest_date_col = 'Fecha'
        dest_essential_headers = ['Contenedor', 'Placa 2', 'Hr salida QP', dest_date_col]
        print(f"Locating header row in destination file: {main_file_path}, sheet 'BBDD'...")
        dest_header_row = find_header_row(main_file_path, 'BBDD', dest_essential_headers)
        
        print(f"Loading main file: {main_file_path} (using header row {dest_header_row + 1})")
        main_df = pd.read_excel(main_file_path, sheet_name='BBDD', header=dest_header_row)
        
        current_sheet_day_str = sheet_date
        match_ym = re.search(r"_(\d{4})_(\d{2})", os.path.basename(export_log_path))
        if not match_ym:
            raise ValueError(f"Could not extract year and month from export log filename: {export_log_path}.")
        export_year = int(match_ym.group(1))
        export_month = int(match_ym.group(2))
        try:
            processing_date = datetime(export_year, export_month, int(current_sheet_day_str)).date()
            print(f"Targeting updates for date: {processing_date.strftime('%Y-%m-%d')}")
        except ValueError as ve:
            raise ValueError(f"Invalid day '{current_sheet_day_str}' for month {export_month}, year {export_year}. Error: {ve}")

        if dest_date_col not in main_df.columns:
            raise ValueError(f"Date column '{dest_date_col}' not found in destination file {main_file_path}.")
        
        main_df[dest_date_col] = pd.to_datetime(main_df[dest_date_col], errors='coerce').dt.date
        if main_df[dest_date_col].isnull().all() and not main_df.empty:
            print(f"WARNING: All dates in column '{dest_date_col}' of destination file are invalid or empty.")
        elif main_df[dest_date_col].isnull().any():
            print(f"WARNING: Some dates in column '{dest_date_col}' of destination file could not be parsed.")

        src_essential_headers = ['NUMERO CONTENEDOR', 'PLACA DE CARRETA', 'HORA DE SALIDA']
        try:
            print(f"Locating header row in source file: {export_log_path}, sheet '{sheet_date}'...")
            header_row_src = find_header_row(export_log_path, sheet_date, src_essential_headers)
            export_df = pd.read_excel(export_log_path, sheet_name=sheet_date, header=header_row_src)
        except Exception as e:
            print(f"Error loading sheet '{sheet_date}': {str(e)}. Trying to find fallback sheet...")
            wb = load_workbook(export_log_path, read_only=True)
            sheets = wb.sheetnames
            found_sheet = None
            if sheet_date in sheets:
                found_sheet = sheet_date
            else:
                for s_name in sheets:
                    if isinstance(s_name, str) and sheet_date in s_name:
                        found_sheet = s_name
                        break
            if found_sheet:
                print(f"Found matching sheet: {found_sheet}")
                header_row_src_fallback = find_header_row(export_log_path, found_sheet, src_essential_headers)
                export_df = pd.read_excel(export_log_path, sheet_name=found_sheet, header=header_row_src_fallback)
            else:
                raise ValueError(f"Could not find sheet for day '{sheet_date}'. Available: {sheets}")
        
        print(f"\nMain file (Destination) columns: {main_df.columns.tolist()}")
        print(f"Export file (Source) columns: {export_df.columns.tolist()}")
        
        dest_container_col = 'Contenedor'
        dest_plate_col = 'Placa 2'
        dest_time_col = 'Hr salida QP'
        required_dest_cols = [dest_container_col, dest_plate_col, dest_time_col, dest_date_col]
        if any(col not in main_df.columns for col in required_dest_cols):
            raise ValueError(f"Missing required columns in destination: {main_file_path}")

        export_df.columns = [str(col).strip() for col in export_df.columns]
        expected_source_container = 'NUMERO CONTENEDOR'
        expected_source_plate = 'PLACA DE CARRETA'
        expected_source_time = 'HORA DE SALIDA'
        source_container_col = expected_source_container if expected_source_container in export_df.columns else next((c for c in export_df.columns if 'CONTENEDOR' in c.upper() or 'CONTAINER' in c.upper()), None)
        source_plate_col = expected_source_plate if expected_source_plate in export_df.columns else next((c for c in export_df.columns if ('PLACA' in c.upper() and 'CARRETA' in c.upper())), None)
        source_time_col = expected_source_time if expected_source_time in export_df.columns else next((c for c in export_df.columns if 'HORA' in c.upper() and 'SALIDA' in c.upper()), None)

        if not all([source_container_col, source_plate_col, source_time_col]):
            raise ValueError("Could not identify all source columns.")
        print(f"Identified Source Columns: C='{source_container_col}', P='{source_plate_col}', T='{source_time_col}'")
        print(f"Identified Dest Columns: C='{dest_container_col}', P='{dest_plate_col}', T='{dest_time_col}'")

        main_df[dest_container_col] = main_df[dest_container_col].astype(str).fillna('')
        main_df[dest_plate_col] = main_df[dest_plate_col].astype(str).fillna('')
        export_df[source_container_col] = export_df[source_container_col].astype(str).fillna('')
        export_df[source_plate_col] = export_df[source_plate_col].astype(str).fillna('')
        
        main_df['Contenedor_norm'] = main_df[dest_container_col].str.replace(r'[- ]', '', regex=True).str.upper()
        main_df['Placa_norm'] = main_df[dest_plate_col].str.replace(r'[- ]', '', regex=True).str.upper()
        export_df['Container_norm'] = export_df[source_container_col].str.replace(r'[- ]', '', regex=True).str.upper()
        export_df['Plate_norm'] = export_df[source_plate_col].str.replace(r'[- ]', '', regex=True).str.upper()
        
        print("\nSample normalized dest values:")
        print(main_df[['Contenedor_norm', 'Placa_norm']].head())
        print("\nSample normalized source values:")
        print(export_df[['Container_norm', 'Plate_norm']].head())
        
        departure_times = {}
        for _, row in export_df.iterrows():
            container, plate = row['Container_norm'], row['Plate_norm']
            time_val = row[source_time_col]
            if container and plate and pd.notna(time_val):
                fmt_time = None
                if isinstance(time_val, (time, datetime)): fmt_time = time_val.strftime('%H:%M')
                elif isinstance(time_val, str):
                    m = re.search(r'(\d{1,2}):(\d{2})', time_val)
                    if m: fmt_time = f"{int(m.group(1)):02d}:{m.group(2)}"
                elif isinstance(time_val, (int, float)) and 0 <= time_val < 1:
                    secs = int(time_val * 86400)
                    fmt_time = f"{secs // 3600:02d}:{(secs % 3600) // 60:02d}"
                else:
                    try: fmt_time = pd.to_datetime(str(time_val), errors='coerce').time().strftime('%H:%M')
                    except: pass
                if fmt_time: departure_times[(container, plate)] = fmt_time

        print(f"\nMatching... Dest rows: {len(main_df)}, Src lookup: {len(departure_times)}")
        for idx, row in main_df.iterrows():
            key = (row['Contenedor_norm'], row['Placa_norm'])
            if key in departure_times:
                current_val, new_time = main_df.at[idx, dest_time_col], departure_times[key]
                curr_time_str = None
                if isinstance(current_val, (time, datetime)): curr_time_str = current_val.strftime('%H:%M')
                elif pd.notna(current_val): curr_time_str = str(current_val)
                
                if pd.isna(current_val) or curr_time_str != new_time:
                    dest_row_date = main_df.at[idx, dest_date_col] 
                    if pd.notna(dest_row_date) and dest_row_date == processing_date:
                        main_df.at[idx, dest_time_col] = new_time
                        updates_count += 1
        
        main_df = main_df.drop(['Contenedor_norm', 'Placa_norm'], axis=1)
        
        if updates_count > 0:
            try:
                print(f"\nSaving updates directly to original file (preserving formatting and structure): {main_file_path}, sheet: BBDD")
                
                book = load_workbook(main_file_path)
                if 'BBDD' not in book.sheetnames:
                    raise ValueError(f"Sheet 'BBDD' not found in workbook '{main_file_path}'.")
                sheet = book['BBDD']
                
                # Convert DataFrame to rows, including header
                rows_to_write = dataframe_to_rows(main_df, index=False, header=True)

                # dest_header_row is the 0-indexed row where headers were FOUND.
                # We start writing new headers (from main_df) at this same 0-indexed row.
                # openpyxl cell row indices are 1-based.
                start_excel_row_for_writing = dest_header_row + 1 

                for r_idx_offset, row_values in enumerate(rows_to_write):
                    current_excel_row = start_excel_row_for_writing + r_idx_offset
                    for c_idx_offset, value in enumerate(row_values):
                        current_excel_col = c_idx_offset + 1 # openpyxl columns are 1-based
                        sheet.cell(row=current_excel_row, column=current_excel_col, value=value)
                
                book.save(main_file_path)
                
                print(f"Successfully updated sheet 'BBDD' in {main_file_path} using openpyxl direct cell writing.")
                print(f"\nUpdated {updates_count} entries in the original file.")
                return updates_count, main_file_path

            except Exception as e_save:
                raise ValueError(f"Error saving updated file using openpyxl: {e_save}")
        else:
            print("No matching entries found to update, or values already matched. The file was not modified.")
            return 0, main_file_path 

    except ValueError as ve:
        print(f"ERROR: {ve}")
        return 0, None
    except Exception as e:
        import traceback
        print(f"ERROR: {e}")
        print(traceback.format_exc())
        return 0, None

def main():
    print("=== INICIANDO SCRIPT DE ACTUALIZACIÓN AUTOMÁTICA ===")
    yesterday = datetime.now() - timedelta(days=1)
    yesterday_day_str = yesterday.strftime('%d') # Day as string, e.g., '05', '14'
    yesterday_year_str = yesterday.strftime('%Y') # Year as string, e.g., '2025'
    yesterday_month_str = yesterday.strftime('%m') # Month as string, e.g., '01', '12'

    # --- Configuration - Adjust these paths for your server environment --- 
    # Make sure to use raw strings (r"...") or double backslashes (\\) for Windows paths.
    destination_file_path = "C:\\Users\\jrivas\\Documents\\Leadtimes\\Lead time OQ1 Exportaciones 2025.xlsx"
    source_files_directory = "C:\\Users\\jrivas\\Documents\\Leadtimes"
    source_file_name_template = "SALIDA_{year}_{month}.xlsx"
    # --- End Configuration ---

    source_file_name = source_file_name_template.format(year=yesterday_year_str, month=yesterday_month_str)
    source_file_path = os.path.join(source_files_directory, source_file_name)

    print(f"Fecha de procesamiento (ayer): {yesterday.strftime('%Y-%m-%d')}")
    print(f"Archivo DESTINO: {destination_file_path}")
    print(f"Archivo ORIGEN (esperado): {source_file_path}")
    print(f"Día para la hoja del ORIGEN: {yesterday_day_str}")

    if not os.path.exists(destination_file_path):
        print(f"ERROR: El archivo DESTINO '{destination_file_path}' no existe.")
        return

    if not os.path.exists(source_file_path):
        print(f"ERROR: El archivo ORIGEN '{source_file_path}' para la fecha de ayer no existe.")
        print("Por favor, asegúrese de que el archivo con el formato correcto (ej: SALIDA_YYYY_MM.xlsx) esté en la carpeta sincronizada.")
        return

    try:
        print("\nIniciando la función de actualización...")
        updated_count, saved_to_path = update_export_times(
            main_file_path=destination_file_path, 
            export_log_path=source_file_path, 
            sheet_date=yesterday_day_str
        )
        print("\n=== RESUMEN DE OPERACIÓN ===")
        if updated_count > 0:
            print(f"* Se actualizaron exitosamente {updated_count} horas de salida en la columna 'Hr salida QP'.")
            print(f"* El archivo de destino '{saved_to_path}' fue actualizado directamente.")
            print(f"(La hoja 'BBDD' en '{saved_to_path}' fue actualizada.)")
        elif updated_count == 0:
            print("- No se realizaron actualizaciones. Esto puede ser normal si no había datos nuevos o coincidencias.")
            print(f"  El archivo de destino NO fue modificado: {destination_file_path}")

    except FileNotFoundError as fnf_error:
        print(f"\nERROR CRÍTICO - Archivo no encontrado: {fnf_error}")
        print("Por favor, verifique que las rutas y nombres de archivo sean correctos y que los archivos existan.")
    except ValueError as val_error:
        print(f"\nERROR CRÍTICO - Error de datos o configuración: {val_error}")
        print("Esto puede deberse a columnas faltantes, encabezados incorrectos, fechas inválidas, o problemas con el formato del nombre del archivo de origen.")
    except Exception as e:
        print(f"\nERROR CRÍTICO - Ocurrió un error inesperado: {e}")
        import traceback
        traceback.print_exc() # Prints detailed traceback for unexpected errors
    finally:
        print("\n=== SCRIPT DE ACTUALIZACIÓN AUTOMÁTICA FINALIZADO ===")

if __name__ == "__main__":
    main()
