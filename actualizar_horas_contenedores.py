import pandas as pd
import os
from datetime import datetime, timedelta, time
import re
from openpyxl import load_workbook

def find_header_row(file_path, sheet_name, essential_headers, max_rows_to_check=20):
    """
    Finds the 0-indexed row number containing all essential_headers.
    Searches within the first `max_rows_to_check` of the specified sheet.
    Returns the 0-indexed row number if found, otherwise raises ValueError.
    """
    try:
        df_peek = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=max_rows_to_check)
    except Exception as e:
        # Try with openpyxl engine if default fails, as it can be more robust for varied files
        try:
            df_peek = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=max_rows_to_check, engine='openpyxl')
        except Exception as e_openpyxl:
            raise ValueError(f"Could not read sheet '{sheet_name}' from '{file_path}' to find headers. Default engine error: {e}. Openpyxl engine error: {e_openpyxl}")

    normalized_essential_headers = {str(h).strip().lower() for h in essential_headers}
    if not normalized_essential_headers:
        raise ValueError("Essential headers list cannot be empty.")

    for i, row in df_peek.iterrows():
        row_values = {str(cell).strip().lower() for cell in row.tolist() if pd.notna(cell) and str(cell).strip()}
        if normalized_essential_headers.issubset(row_values):
            print(f"Found all essential headers ({', '.join(essential_headers)}) in row {i+1} (0-indexed: {i}) of sheet '{sheet_name}'.")
            return i 

    raise ValueError(
        f"Could not find all essential headers ({', '.join(essential_headers)}) "
        f"in the first {max_rows_to_check} rows of sheet '{sheet_name}' in file '{file_path}'. "
        f"Please ensure these columns exist and are within the first {max_rows_to_check} rows."
    )

def update_export_times(main_file_path, export_log_path, sheet_date=None):
    """
    Updates the 'Hr Salida QP' column in the main file with times from the export container log
    by matching container numbers and license plates.
    
    Parameters:
    -----------
    main_file_path : str
        Path to the main Excel file (e.g., Lead time OQ1 Exportaciones 2025.xlsx)
    export_log_path : str
        Path to the export container log Excel file (e.g., SALIDA_2025_05.xlsx)
    sheet_date : str, optional
        Date in format 'DD' representing the sheet name to load in the export log file
        If None, will use the PREVIOUS day of the month (not current day)
    """
    updates_count = 0 # <<< INICIALIZAR CONTADOR
    
    # Determine the sheet name to use in the export log file
    if sheet_date is None:
        # CAMBIO AQUÍ: Usar el día anterior en lugar del día actual
        yesterday = datetime.now() - timedelta(days=1)
        sheet_date = yesterday.strftime('%d')
    
    print(f"Loading export log data from sheet: {sheet_date} (Previous day by default)")
    
    try:
        # Define essential headers for destination file
        dest_date_col = 'Fecha' # <--- ADDED Destination date column name
        dest_essential_headers = ['Contenedor', 'Placa 2', 'Hr salida QP', dest_date_col] # <--- ADDED dest_date_col
        print(f"Locating header row in destination file: {main_file_path}, sheet 'BBDD'...")
        header_row_dest = find_header_row(main_file_path, 'BBDD', dest_essential_headers)
        
        print(f"Loading main file: {main_file_path} (using header row {header_row_dest + 1})")
        main_df = pd.read_excel(main_file_path, sheet_name='BBDD', header=header_row_dest)
        
        # --- Determine Processing Date ---
        current_sheet_day_str = sheet_date
        if current_sheet_day_str is None:
            # This logic should align with how sheet_date is determined if None (e.g., yesterday)
            # The script's main function sets sheet_date to yesterday's day string if input is empty
            # So, sheet_date entering this function should already be a day string.
            # For safety, if it were None here, we'd need to re-evaluate or raise error.
            # Assuming sheet_date is always a day string by this point.
            effective_date_for_processing = datetime.now() - timedelta(days=1)
            current_sheet_day_str = effective_date_for_processing.strftime('%d')
        
        match_ym = re.search(r"_(\d{4})_(\d{2})", os.path.basename(export_log_path))
        if not match_ym:
            raise ValueError(f"Could not extract year and month from export log filename: {export_log_path}. Expected format like 'NAME_YYYY_MM.xlsx' or 'YYYY_MM_NAME.xlsx'")
        export_year = int(match_ym.group(1))
        export_month = int(match_ym.group(2))
        try:
            processing_date = datetime(export_year, export_month, int(current_sheet_day_str)).date()
            print(f"Targeting updates for date: {processing_date.strftime('%Y-%m-%d')}")
        except ValueError as ve:
            raise ValueError(f"Invalid day '{current_sheet_day_str}' for month {export_month}, year {export_year} derived from source file. Error: {ve}")

        # Convert destination date column to datetime.date objects
        if dest_date_col not in main_df.columns:
             # This should be caught by find_header_row if dest_date_col is in dest_essential_headers
            raise ValueError(f"Date column '{dest_date_col}' not found in destination file {main_file_path}.")
        
        main_df[dest_date_col] = pd.to_datetime(main_df[dest_date_col], errors='coerce')
        if main_df[dest_date_col].isnull().all() and not main_df.empty:
            print(f"WARNING: All dates in column '{dest_date_col}' of destination file are invalid or empty after conversion. Date matching will likely fail.")
        elif main_df[dest_date_col].isnull().any():
            print(f"WARNING: Some dates in column '{dest_date_col}' of destination file could not be parsed. These rows won't match by date.")
        main_df[dest_date_col] = main_df[dest_date_col].dt.date # Convert to date objects for comparison

        # Define essential headers for source file
        src_essential_headers = ['NUMERO CONTENEDOR', 'PLACA DE CARRETA', 'HORA DE SALIDA']

        # Load the export log file (source) from the specified sheet
        try:
            print(f"Locating header row in source file: {export_log_path}, sheet '{sheet_date}'...")
            header_row_src = find_header_row(export_log_path, sheet_date, src_essential_headers)
            print(f"Attempting to load export log from sheet: {sheet_date} (using header row {header_row_src + 1})")
            export_df = pd.read_excel(export_log_path, sheet_name=sheet_date, header=header_row_src)
        except Exception as e:
            print(f"Error loading sheet '{sheet_date}' or finding its headers: {str(e)}")
            print("Attempting to find the sheet with previous day's data...")
            
            # Try to find a sheet with previous day's date in different formats
            try:
                wb = load_workbook(export_log_path, read_only=True)
                sheets = wb.sheetnames
            except Exception as load_err:
                 raise ValueError(f"Could not open or read sheets from export file {export_log_path}: {load_err}")

            print(f"Available sheets: {sheets}")
            
            yesterday_date_str = sheet_date  # Ya está calculado arriba
            
            found_sheet = None
            if yesterday_date_str in sheets:
                found_sheet = yesterday_date_str
            else:
                 # If sheet 'DD' not found, check if any sheet *contains* the day number (e.g. 'Sheet 12')
                 for sheet in sheets:
                     if isinstance(sheet, str) and yesterday_date_str in sheet:
                         found_sheet = sheet
                         break
            
            if found_sheet:
                print(f"Found matching sheet: {found_sheet}")
                print(f"Locating header row in source file: {export_log_path}, sheet '{found_sheet}'...")
                header_row_src_fallback = find_header_row(export_log_path, found_sheet, src_essential_headers)
                print(f"Attempting to load export log from sheet: {found_sheet} (using header row {header_row_src_fallback + 1})")
                export_df = pd.read_excel(export_log_path, sheet_name=found_sheet, header=header_row_src_fallback)
            else:
                raise ValueError(f"Could not find sheet for day '{yesterday_date_str}'. Available sheets: {sheets}")
        
        # Print column names for debugging
        print(f"\nMain file (Destination) columns: {main_df.columns.tolist()}")
        print(f"Export file (Source) columns: {export_df.columns.tolist()}")
        
        # --- Column Identification --- 
        # Destination File Columns (Main File)
        dest_container_col = 'Contenedor' # Columna X
        dest_plate_col = 'Placa 2'       # Columna Z
        dest_time_col = 'Hr salida QP'   # Columna AB
        # dest_date_col is already defined above
        
        required_dest_cols = [dest_container_col, dest_plate_col, dest_time_col, dest_date_col] # <--- ADDED dest_date_col
        missing_dest_cols = [col for col in required_dest_cols if col not in main_df.columns]
        if missing_dest_cols:
            raise ValueError(f"Missing required columns in destination file ({main_file_path}): {missing_dest_cols}")

        # Source File Columns (Export Log)
        export_df.columns = [col.strip() if isinstance(col, str) else col for col in export_df.columns]
        
        # Identify source columns dynamically but prioritize expected names
        source_container_col = None
        source_plate_col = None
        source_time_col = None

        # Prioritize expected names from screenshots
        expected_source_container = 'NUMERO CONTENEDOR'
        expected_source_plate = 'PLACA DE CARRETA'
        expected_source_time = 'HORA DE SALIDA'

        if expected_source_container in export_df.columns:
            source_container_col = expected_source_container
        if expected_source_plate in export_df.columns:
            source_plate_col = expected_source_plate
        if expected_source_time in export_df.columns:
            source_time_col = expected_source_time

        # Fallback to searching if exact names aren't found
        if not source_container_col:
            for col in export_df.columns:
                 if isinstance(col, str) and ('CONTENEDOR' in col.upper() or 'CONTAINER' in col.upper() or 'NUMERO CONTENEDO' in col.upper()):
                     source_container_col = col
                     print(f"Found source container column (fallback): '{col}'")
                     break
        if not source_plate_col:
            for col in export_df.columns:
                 if isinstance(col, str) and (('PLACA' in col.upper() and 'CARRETA' in col.upper()) or 
                                              ('PLACA DE CARRET' in col.upper())):
                     source_plate_col = col
                     print(f"Found source plate column (fallback): '{col}'")
                     break
        if not source_time_col:
             for col in export_df.columns:
                 if isinstance(col, str) and ('HORA' in col.upper() and 'SALIDA' in col.upper()):
                     source_time_col = col
                     print(f"Found source time column (fallback): '{col}'")
                     break

        # Raise errors if columns couldn't be identified
        if not source_container_col:
            raise ValueError(f"Could not find Container column in source file ({export_log_path}). Looked for '{expected_source_container}' or similar.")
        if not source_plate_col:
            raise ValueError(f"Could not find Plate column in source file ({export_log_path}). Looked for '{expected_source_plate}' or similar.")
        if not source_time_col:
            raise ValueError(f"Could not find Time column in source file ({export_log_path}). Looked for '{expected_source_time}' or similar.")
        
        print(f"Identified Source Columns: Container='{source_container_col}', Plate='{source_plate_col}', Time='{source_time_col}'")
        print(f"Identified Destination Columns: Container='{dest_container_col}', Plate='{dest_plate_col}', Time='{dest_time_col}'")

        # --- Data Normalization --- 
        # Convert to string and handle potential errors
        main_df[dest_container_col] = main_df[dest_container_col].astype(str).fillna('')
        main_df[dest_plate_col] = main_df[dest_plate_col].astype(str).fillna('')
        export_df[source_container_col] = export_df[source_container_col].astype(str).fillna('')
        export_df[source_plate_col] = export_df[source_plate_col].astype(str).fillna('')
        
        # Normalize: remove spaces, hyphens, convert to uppercase
        main_df['Contenedor_norm'] = main_df[dest_container_col].str.strip().str.replace(r'[- ]', '', regex=True).str.upper()
        main_df['Placa_norm'] = main_df[dest_plate_col].str.strip().str.replace(r'[- ]', '', regex=True).str.upper()
        export_df['Container_norm'] = export_df[source_container_col].str.strip().str.replace(r'[- ]', '', regex=True).str.upper()
        export_df['Plate_norm'] = export_df[source_plate_col].str.strip().str.replace(r'[- ]', '', regex=True).str.upper()
        
        # Print some debug info
        print("\nSample normalized values from destination file:")
        print(main_df[['Contenedor_norm', 'Placa_norm']].head())
        print("\nSample normalized values from source file:")
        print(export_df[['Container_norm', 'Plate_norm']].head())
        
        # --- Create Lookup Dictionary --- 
        # Store exact container/plate pairs and their departure times from the source file
        departure_times = {}
        for _, row in export_df.iterrows():
            container = row['Container_norm']
            plate = row['Plate_norm']
            departure_time_val = row[source_time_col]
            
            # Only consider non-empty container/plate and valid times
            if container and plate and pd.notna(departure_time_val):
                # Format departure time as string HH:MM
                formatted_time = None
                if isinstance(departure_time_val, time): # Check if it's a time object
                    formatted_time = departure_time_val.strftime('%H:%M') 
                elif isinstance(departure_time_val, datetime): # Check if it's a datetime object
                    formatted_time = departure_time_val.strftime('%H:%M')
                elif isinstance(departure_time_val, str): # Check if it's a string
                    time_match = re.search(r'(\d{1,2}):(\d{2})', departure_time_val)
                    if time_match:
                        hours, minutes = time_match.groups()
                        formatted_time = f"{int(hours):02d}:{minutes}"
                else: # Try converting other types like numbers (Excel time representation)
                    try:
                        # Handle Excel time serial numbers (float between 0 and 1)
                        if isinstance(departure_time_val, (int, float)) and 0 <= departure_time_val < 1:
                            total_seconds = int(departure_time_val * 24 * 60 * 60)
                            hours = total_seconds // 3600
                            minutes = (total_seconds % 3600) // 60
                            formatted_time = f"{hours:02d}:{minutes:02d}"
                        else:
                             # Attempt direct conversion to time if possible
                             temp_time = pd.to_datetime(str(departure_time_val), errors='coerce').time()
                             if pd.notna(temp_time):
                                  formatted_time = temp_time.strftime('%H:%M')
                    except Exception:
                         pass # Ignore if conversion fails

                if formatted_time: # Only store if time was successfully formatted
                     key = (container, plate)
                     # Store the latest time if duplicates exist
                     departure_times[key] = formatted_time 

        # --- Matching and Updating --- 
        print(f"\nMatching process starting...")
        print(f"Total rows in destination file: {len(main_df)}")
        print(f"Total unique Container/Plate pairs in source file lookup: {len(departure_times)}")
        
        # Iterate through the destination file and update based on exact match
        for idx, row in main_df.iterrows():
            container = row['Contenedor_norm']
            plate = row['Placa_norm']
            
            key = (container, plate)
            
            # <<< CAMBIO: Check for exact match ONLY
            if key in departure_times:
                current_value = main_df.at[idx, dest_time_col]
                new_time = departure_times[key]
                # Update only if the cell is empty or the time is different
                # Convert current value to HH:MM string for comparison if it's time/datetime
                current_time_str = None
                if isinstance(current_value, time): # Corrected: was datetime.time
                    current_time_str = current_value.strftime('%H:%M')
                elif isinstance(current_value, datetime):
                     current_time_str = current_value.strftime('%H:%M')
                elif pd.isna(current_value):
                     current_time_str = None
                else: # Keep as is if it's already a string or other type
                     current_time_str = str(current_value)

                if pd.isna(current_value) or current_time_str != new_time:
                    destination_row_date = main_df.loc[idx, dest_date_col]
                    if pd.isna(destination_row_date):
                        # print(f"Debug: Row index {idx} in destination has NaT/None date. Skipping date check for this row.")
                        continue # Skip if date in destination is unparseable
                    
                    if destination_row_date == processing_date:
                        main_df.at[idx, dest_time_col] = new_time
                        updates_count += 1
                        # print(f"Match found for {container}, {plate}: Updating '{dest_time_col}' to {new_time}") # Optional: uncomment for verbose logging
        
        # Drop temporary normalization columns
        main_df = main_df.drop(['Contenedor_norm', 'Placa_norm'], axis=1)
        
        # --- Save Output --- 
        output_path = main_file_path.replace('.xlsx', '_updated.xlsx')
        print(f"\nPreparing to save updated file to: {output_path}")
        
        try:
            # Use ExcelWriter to preserve existing sheets if possible, but focus on writing the updated data
            with pd.ExcelWriter(output_path, engine='openpyxl', mode='w') as writer: # mode='w' overwrites
                main_df.to_excel(writer, sheet_name='BBDD', index=False)
            print(f"Successfully saved updated data to sheet 'BBDD' in {output_path}")
        except Exception as save_err:
             print(f"\nError saving the file: {save_err}")
             print("Attempting to save as a new file instead...")
             try:
                 output_path_new = main_file_path.replace('.xlsx', '_updated_new.xlsx')
                 main_df.to_excel(output_path_new, sheet_name='BBDD', index=False)
                 print(f"Successfully saved updated data to a new file: {output_path_new}")
                 output_path = output_path_new # Update output path variable
             except Exception as fallback_save_err:
                 raise IOError(f"Could not save the updated file. Error: {fallback_save_err}")

        print(f"\nUpdated {updates_count} entries in the destination file.")
        
        return updates_count, output_path
    
    except FileNotFoundError as fnf_error:
        print(f"\n❌ ERROR: File not found. {fnf_error}")
        print("Please ensure the file paths entered are correct and the files exist.")
        return 0, None
    except ValueError as val_error:
         print(f"\n❌ ERROR: {val_error}")
         print("Please check the column names and sheet names in your Excel files.")
         return 0, None
    except Exception as e:
        import traceback
        print(f"\n❌ An unexpected error occurred: {str(e)}")
        print(traceback.format_exc()) # Print detailed traceback for debugging
        return 0, None

def main():
    """Main function to run the export times updater"""
    print("\n===== ACTUALIZADOR DE HORAS DE SALIDA DE CONTENEDORES =====\n")
    print("Este script actualiza la columna 'Hr salida QP' en el archivo principal")
    print("utilizando datos del archivo de registro de salidas de exportación.")
    print("IMPORTANTE: Descargue los archivos de SharePoint/OneDrive a su PC primero.")
    print("Luego, ingrese las rutas locales a los archivos descargados.\n")
    print("NOTA: Por defecto, el programa usará los datos del DÍA ANTERIOR.")
    
    # Get local file paths from user
    main_file_path = input("Ingrese la RUTA LOCAL al archivo Excel DESTINO (Lead time OQ1 Exportaciones 2025): ").strip().strip('"')
    export_log_path = input("Ingrese la RUTA LOCAL al archivo Excel ORIGEN (SALIDA_2025_05): ").strip().strip('"')
    
    # Optional: Specify the sheet date (day of month)
    yesterday = datetime.now() - timedelta(days=1)
    yesterday_str = yesterday.strftime('%d')
    
    sheet_date_input = input(f"Ingrese el NÚMERO del día para la hoja del archivo ORIGEN (Por defecto: '{yesterday_str}' - día anterior): ")
    sheet_date = sheet_date_input.strip() if sheet_date_input.strip() else None
    
    # Basic validation of paths
    if not os.path.exists(main_file_path):
        print(f"\n❌ ERROR: El archivo destino no se encontró en: {main_file_path}")
        input("\nPresione Enter para salir...")
        return
    if not os.path.exists(export_log_path):
        print(f"\n❌ ERROR: El archivo origen no se encontró en: {export_log_path}")
        input("\nPresione Enter para salir...")
        return
        
    print("\nProcesando archivos...")
    # Run the update process
    update_count, output_path = update_export_times(main_file_path, export_log_path, sheet_date)
    
    # Display results
    if output_path:
        if update_count > 0:
            print("\n=== RESUMEN DE OPERACIÓN ===")
            print(f"✅ Se actualizaron exitosamente {update_count} horas de salida en la columna 'Hr salida QP'.")
            print(f"✅ El archivo actualizado se guardó como: {output_path}")
            print("   (Se actualizó la hoja 'BBDD')")
            print("\nPuede abrir este archivo para verificar los cambios.")
        else:
            print("\n=== RESUMEN DE OPERACIÓN ===")
            print("✅ Proceso completado, pero no se realizaron nuevas actualizaciones.")
            print("   Esto puede ser porque no se encontraron coincidencias exactas (Contenedor + Placa)")
            print("   o porque los valores ya estaban actualizados.")
            print(f"   El archivo de salida se generó igualmente en: {output_path}")
    # else case handled by errors inside update_export_times
            
    input("\nPresione Enter para salir...")

if __name__ == "__main__":
    main()